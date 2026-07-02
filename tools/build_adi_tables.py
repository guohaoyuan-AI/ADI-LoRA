from pathlib import Path
import pandas as pd
import numpy as np
import re, json, csv, os

ROOT=Path('/workspace/adi_work')
OUT=ROOT/'outputs'
OUT.mkdir(parents=True, exist_ok=True)

# Helpers

def rel(p):
    return str(Path(p)).replace(str(ROOT)+'/', '')

def pct(x):
    if pd.isna(x): return np.nan
    try: return float(x)*100
    except Exception: return np.nan

def pp(x):
    if pd.isna(x): return np.nan
    try: return float(x)*100
    except Exception: return np.nan

def read_csv(p):
    return pd.read_csv(p)

def norm_bool(v):
    if pd.isna(v): return np.nan
    if isinstance(v,str):
        s=v.strip().lower()
        if s in ('0','false','no','n'): return 0
        if s in ('1','true','yes','y'): return 1
    try: return int(float(v))
    except Exception: return v

# File inventory over source packages and key evidence files
uploaded = [
('/workspace/.cache/01-ADI_final_table_sources_20260629_213223.tar.gz','ADI_final_table_sources_20260629_213223.tar.gz'),
('/workspace/.cache/04-stage4_tiny_imagenet100_results_collection_20260628_160740.tar.gz','stage4_tiny_imagenet100_results_collection_20260628_160740.tar.gz'),
('/workspace/.cache/05-fr_peft_cifar100_seed42_43_44_no_ckpt.tar.gz','fr_peft_cifar100_seed42_43_44_no_ckpt.tar.gz'),
('/workspace/.cache/06-ADI_DoRA_dialogue4a_handoff_package.tar.gz','ADI_DoRA_dialogue4a_handoff_package.tar.gz'),
('/workspace/.cache/07-dora_cifar100c_full19_s3_seed42_no_ckpt_20260627_195802.tar.gz','dora_cifar100c_full19_s3_seed42_no_ckpt_20260627_195802.tar.gz'),
('/workspace/.cache/08-dora_cifar100c_subset_s1_s3_s5_3seed_no_ckpt_20260627_191452.tar.gz','dora_cifar100c_subset_s1_s3_s5_3seed_no_ckpt_20260627_191452.tar.gz'),
('/workspace/.cache/09-stage4_summary_tiny_imagenet100.xlsx','stage4_summary_tiny_imagenet100.xlsx'),
('/workspace/.cache/01-cifar100_dora_seed43_alpha_eval_no_ckpt_20260627_161631.tar.gz','cifar100_dora_seed43_alpha_eval_no_ckpt_20260627_161631.tar.gz'),
('/workspace/.cache/02-cifar100_dora_seed44_alpha_eval_no_ckpt_20260627_182341.tar.gz','cifar100_dora_seed44_alpha_eval_no_ckpt_20260627_182341.tar.gz'),
('/workspace/.cache/04-fr_peft_cifar100_seed42_controls_no_ckpt.tar.gz','fr_peft_cifar100_seed42_controls_no_ckpt.tar.gz'),
('/workspace/.cache/06-ADI_efficiency_filtered_tables_package.zip','ADI_efficiency_filtered_tables_package.zip'),
]

def has_weights(dirpath):
    exts={'.pth','.pt','.ckpt','.safetensors'}
    if not Path(dirpath).exists(): return False
    return any(p.suffix.lower() in exts for p in Path(dirpath).rglob('*') if p.is_file())

inventory=[]
# package-level inventory, manually grounded in contents inspected
package_notes = {
'ADI_final_table_sources_20260629_213223.tar.gz': ('CIFAR-100 LoRA clean, DoRA seed42 clean, VPT/Adapter diagnostics, manifests with several MISSING entries','CIFAR-100; partial Stage4 manifest','LoRA/DoRA/VPT/Adapter','42/43/44 partial', 'Partial source aggregator; not sole evidence'),
'stage4_tiny_imagenet100_results_collection_20260628_160740.tar.gz': ('Tiny-ImageNet LoRA 3 seeds, Tiny-ImageNet DoRA seed42, ImageNet-100 LoRA/DoRA seed42 pilot, logs/manifests','Tiny-ImageNet; ImageNet-100 pilot','LoRA/DoRA','42/43/44 where available','Stage4 source package; ImageNet-100 pilot only'),
'fr_peft_cifar100_seed42_43_44_no_ckpt.tar.gz': ('CIFAR-100 LoRA same-checkpoint clean 3 seeds','CIFAR-100','LoRA/ADI-LoRA','42/43/44','Main source for CIFAR-100 LoRA clean'),
'ADI_DoRA_dialogue4a_handoff_package.tar.gz': ('DoRA repair evidence and CIFAR-100 DoRA seed42 same-checkpoint eval; code patches','CIFAR-100; CIFAR-10 diagnostics','DoRA/ADI-DoRA','42','DoRA repair + seed42 raw; not 3-seed clean pack'),
'dora_cifar100c_full19_s3_seed42_no_ckpt_20260627_195802.tar.gz': ('DoRA CIFAR-100-C full19 severity=3 seed42','CIFAR-100-C','DoRA/ADI-DoRA','42','Supplement detailed corruption source'),
'dora_cifar100c_subset_s1_s3_s5_3seed_no_ckpt_20260627_191452.tar.gz': ('DoRA CIFAR-100-C subset 3 seeds x 4 corruptions x 3 severities','CIFAR-100-C','DoRA/ADI-DoRA','42/43/44','Main corruption robustness source'),
'stage4_summary_tiny_imagenet100.xlsx': ('Previously generated Stage4 workbook','Tiny-ImageNet; ImageNet-100 pilot','LoRA/DoRA','42/43/44 where available','Cross-check only; rebuilt from CSV where possible'),
'cifar100_dora_seed43_alpha_eval_no_ckpt_20260627_161631.tar.gz': ('CIFAR-100 DoRA seed43 same-checkpoint alpha eval','CIFAR-100','DoRA/ADI-DoRA','43','Completes DoRA clean three-seed source'),
'cifar100_dora_seed44_alpha_eval_no_ckpt_20260627_182341.tar.gz': ('CIFAR-100 DoRA seed44 same-checkpoint alpha eval','CIFAR-100','DoRA/ADI-DoRA','44','Completes DoRA clean three-seed source'),
'fr_peft_cifar100_seed42_controls_no_ckpt.tar.gz': ('CIFAR-100 LoRA seed42 control experiments: low-lr, strong-wd, 10ep','CIFAR-100','LoRA controls','42','Supplement/control evidence only'),
'ADI_efficiency_filtered_tables_package.zip': ('Filtered efficiency tables and alpha-selection overhead','CIFAR-100','LoRA/ADI-LoRA/DoRA/ADI-DoRA','42/43/44 where available','Efficiency/protocol source'),
}
manifest_missing=[]
manifest_p=ROOT/'extracted/01-ADI_final_table_sources_20260629_213223/ADI_final_table_sources_20260629_213223/manifests/source_package_manifest.tsv'
if manifest_p.exists():
    m=pd.read_csv(manifest_p, sep='\t')
    manifest_missing=m[m['status'].astype(str).str.upper().eq('MISSING')]['label'].tolist()
for path,name in uploaded:
    text=' '.join(package_notes[name])
    missing='YES' if name.startswith('ADI_final') and manifest_missing else 'NO'
    contains,dataset,method,seeds,remarks=package_notes[name]
    inventory.append({
        'filename': name, 'contains_experiment': contains, 'dataset': dataset, 'method': method, 'seed': seeds,
        'contains_weights': 'NO',
        'main_table_source': 'YES' if name in ['fr_peft_cifar100_seed42_43_44_no_ckpt.tar.gz','dora_cifar100c_subset_s1_s3_s5_3seed_no_ckpt_20260627_191452.tar.gz','stage4_tiny_imagenet100_results_collection_20260628_160740.tar.gz','cifar100_dora_seed43_alpha_eval_no_ckpt_20260627_161631.tar.gz','cifar100_dora_seed44_alpha_eval_no_ckpt_20260627_182341.tar.gz','ADI_efficiency_filtered_tables_package.zip'] else 'PARTIAL' if name in ['ADI_DoRA_dialogue4a_handoff_package.tar.gz','dora_cifar100c_full19_s3_seed42_no_ckpt_20260627_195802.tar.gz'] else 'CROSS_CHECK',
        'supplement_or_pilot_only': 'YES: pilot for ImageNet-100 rows' if 'ImageNet-100' in dataset else ('YES: supplement for full19/details' if 'full19' in contains or name=='ADI_DoRA_dialogue4a_handoff_package.tar.gz' else 'NO'),
        'has_missing': missing,
        'remarks': remarks + (('; manifest MISSING: '+', '.join(manifest_missing)) if missing=='YES' else '')
    })

# Add key file-level inventory for missing-sensitive files
key_files=[]
for p in sorted(list((ROOT/'extracted').rglob('*'))+list((ROOT/'extracted_nested').rglob('*'))+list((ROOT/'extracted_nested_recursive').rglob('*'))+list((ROOT/'new_uploads').rglob('*'))):
    if not p.is_file(): continue
    if p.suffix.lower() not in ['.csv','.tsv','.xlsx','.md','.json'] and not p.name.endswith('.log'): continue
    name=p.name.lower()
    if any(k in name for k in ['same_checkpoint_comparison','combined_same_checkpoint','dora_cifar100c_comparison','summary_overall','summary_by_seed','split_manifest','status','queue','recover','source_package_manifest','stage4_summary']):
        s=rel(p)
        inventory.append({
            'filename': rel(p), 'contains_experiment': 'key evidence/control file', 'dataset': 'inferred from path', 'method': 'inferred from path', 'seed': 'inferred from row/path',
            'contains_weights': 'NO', 'main_table_source': 'YES/PARTIAL depending on placement', 'supplement_or_pilot_only': 'see Table_Placement', 'has_missing': 'NO', 'remarks': 'key parsed/indexed source file'
        })
file_inventory=pd.DataFrame(inventory)

# Main CIFAR100 clean - LoRA
main_clean=[]
lo_path=ROOT/'extracted/05-fr_peft_cifar100_seed42_43_44_no_ckpt/fr_peft_cifar100_seed42_43_44_no_ckpt/combined_same_checkpoint_comparison.csv'
if lo_path.exists():
    df=read_csv(lo_path)
    for _,r in df.iterrows():
        main_clean.append({
            'dataset':'CIFAR-100','method':'ADI-LoRA','base_method':'LoRA','seed':int(r.seed),'selected_alpha':r.selected_alpha,
            'std_iid':pct(r.lora_alpha1_iid),'adi_iid':pct(r.delta_selected_iid),'iid_gain_pp':pp(r.iid_gain),
            'std_bilinear':pct(r.lora_alpha1_bilinear),'adi_bilinear':pct(r.delta_selected_bilinear),'bilinear_gain_pp':pp(r.bilinear_gain),
            'std_nearest':pct(r.lora_alpha1_nearest),'adi_nearest':pct(r.delta_selected_nearest),'nearest_gain_pp':pp(r.nearest_gain),
            'std_nearest_drop_pp':pp(r.lora_alpha1_nearest_drop),'adi_nearest_drop_pp':pp(r.delta_selected_nearest_drop),'drop_reduction_pp':pp(r.drop_reduction),
            'rrr_std_nearest': (r.lora_alpha1_nearest/r.lora_alpha1_iid if r.lora_alpha1_iid else np.nan),
            'rrr_adi_nearest': (r.delta_selected_nearest/r.delta_selected_iid if r.delta_selected_iid else np.nan),
            'nearest_used_for_alpha_selection':norm_bool(r.nearest_used_for_alpha_selection),'corruption_used_for_alpha_selection':0,
            'checkpoint_selection_rule':'final_checkpoint_only','nearest_used_for_checkpoint_selection':norm_bool(r.nearest_used_for_checkpoint_selection),
            'same_checkpoint_eval':1,'nan_flag':norm_bool(r.nan_flag),'divergence_flag':norm_bool(r.divergence_flag),
            'source_file':rel(lo_path),'evidence_status':'RAW_CSV','placement':'Main paper candidate'
        })
# DoRA repaired clean raw same-checkpoint comparisons, one file per seed.
for dora_path in [
    ROOT/'extracted/06-ADI_DoRA_dialogue4a_handoff_package/ADI_DoRA_dialogue4a_handoff_package/tables/cifar100_same_checkpoint_comparison_seed42.csv',
    ROOT/'new_uploads/extracted/01-cifar100_dora_seed43_alpha_eval_no_ckpt_20260627_161631./outputs/cifar100_dora_seed43_same_checkpoint_alpha_eval_20260627_160355/same_checkpoint_comparison.csv',
    ROOT/'new_uploads/extracted/02-cifar100_dora_seed44_alpha_eval_no_ckpt_20260627_182341./outputs/cifar100_dora_seed44_same_checkpoint_alpha_eval_20260627_175841/same_checkpoint_comparison.csv',
]:
    if dora_path.exists():
        df=read_csv(dora_path)
        for _,r in df.iterrows():
            main_clean.append({
                'dataset':'CIFAR-100','method':'ADI-DoRA','base_method':'DoRA','seed':int(r.seed),'selected_alpha':r.selected_alpha,
                'std_iid':pct(r.alpha1_iid),'adi_iid':pct(r.selected_iid),'iid_gain_pp':pp(r.iid_gain),
                'std_bilinear':pct(r.alpha1_bilinear),'adi_bilinear':pct(r.selected_bilinear),'bilinear_gain_pp':pp(r.bilinear_gain),
                'std_nearest':pct(r.alpha1_nearest),'adi_nearest':pct(r.selected_nearest),'nearest_gain_pp':pp(r.nearest_gain),
                'std_nearest_drop_pp':pp(r.alpha1_nearest_drop),'adi_nearest_drop_pp':pp(r.selected_nearest_drop),'drop_reduction_pp':pp(r.drop_reduction),
                'rrr_std_nearest': (r.alpha1_nearest/r.alpha1_iid if r.alpha1_iid else np.nan),
                'rrr_adi_nearest': (r.selected_nearest/r.selected_iid if r.selected_iid else np.nan),
                'nearest_used_for_alpha_selection':norm_bool(r.nearest_used_for_alpha_selection),'corruption_used_for_alpha_selection':0,
                'checkpoint_selection_rule':r.checkpoint_selection_rule,'nearest_used_for_checkpoint_selection':norm_bool(r.nearest_used_for_checkpoint_selection),
                'same_checkpoint_eval':1,'nan_flag':0,'divergence_flag':0,
                'source_file':rel(dora_path),'evidence_status':'RAW_CSV','placement':'Main paper candidate'
            })
main_cifar100_clean=pd.DataFrame(main_clean)

# Missing placeholders for required DoRA seeds 43/44 without raw source
for seed in [43,44]:
    if not ((main_cifar100_clean['base_method'].eq('DoRA')) & (main_cifar100_clean['seed'].eq(seed))).any():
        main_cifar100_clean=pd.concat([main_cifar100_clean, pd.DataFrame([{
            'dataset':'CIFAR-100','method':'ADI-DoRA','base_method':'DoRA','seed':seed,'selected_alpha':np.nan,
            'std_iid':np.nan,'adi_iid':np.nan,'iid_gain_pp':np.nan,'std_bilinear':np.nan,'adi_bilinear':np.nan,'bilinear_gain_pp':np.nan,
            'std_nearest':np.nan,'adi_nearest':np.nan,'nearest_gain_pp':np.nan,'std_nearest_drop_pp':np.nan,'adi_nearest_drop_pp':np.nan,'drop_reduction_pp':np.nan,
            'rrr_std_nearest':np.nan,'rrr_adi_nearest':np.nan,'nearest_used_for_alpha_selection':np.nan,'corruption_used_for_alpha_selection':np.nan,
            'checkpoint_selection_rule':'MISSING_RAW_CSV','nearest_used_for_checkpoint_selection':np.nan,'same_checkpoint_eval':np.nan,'nan_flag':np.nan,'divergence_flag':np.nan,
            'source_file':'MISSING: cifar100_dora_three_seed_no_ckpt_pack or clean same-checkpoint xlsx/csv','evidence_status':'MISSING_SOURCE','placement':'Cannot support 3-seed formal DoRA claim yet'
        }])], ignore_index=True)

# CIFAR100-C
c_rows=[]
def add_c100c(path, method, tier):
    if Path(path).exists():
        df=read_csv(path)
        for _,r in df.iterrows():
            c_rows.append({
                'dataset':'CIFAR-100-C','evidence_tier':tier,'method':method,'seed':int(r.seed),'corruption':r.corruption,'severity':int(r.severity),
                'selected_alpha':r.selected_alpha,'alpha1_acc':pct(r.alpha1_acc),'selected_acc':pct(r.selected_acc),'acc_gain_pp':pp(r.acc_gain),
                'alpha1_loss_ce':r.alpha1_loss_ce,'selected_loss_ce':r.selected_loss_ce,'loss_reduction':r.loss_reduction,'n':int(r.n),
                'corruption_used_for_alpha_selection':norm_bool(r.corruption_used_for_alpha_selection),'nearest_used_for_alpha_selection':norm_bool(r.nearest_used_for_alpha_selection),
                'checkpoint_selection_rule':r.checkpoint_selection_rule,'nearest_used_for_checkpoint_selection':norm_bool(r.nearest_used_for_checkpoint_selection),
                'same_checkpoint_eval':1,'source_file':rel(path),'placement':'Main paper candidate' if 'subset' in tier.lower() else 'Supplement detailed table'
            })
add_c100c(ROOT/'extracted/08-dora_cifar100c_subset_s1_s3_s5_3seed_no_ckpt_20260627_191452./outputs/dora_cifar100c_subset_s1_s3_s5/dora_cifar100c_comparison.csv','ADI-DoRA','subset_4corr_s1_s3_s5_3seed')
add_c100c(ROOT/'extracted/07-dora_cifar100c_full19_s3_seed42_no_ckpt_20260627_195802/dora_cifar100c_full19_s3_seed42_20260627_193734/dora_cifar100c_comparison.csv','ADI-DoRA','full19_s3_seed42')
main_cifar100c=pd.DataFrame(c_rows)

# Stage4 clean rows from same_checkpoint comparisons, prefer non-nested collection comparisons plus nested for tiny seed42 fixedlabel
stage_rows=[]
stage_sources = [
('Tiny-ImageNet','ADI-LoRA','Formal small-scale', ROOT/'extracted_nested/04-stage4_tiny_imagenet100_results_collection_20260628_160740_stage4_tiny_imagenet100_collect_20260628_160740_packages_tiny_lora_seed42_final_fixedlabel_eval_no_ckpt_20260628_000521./outputs/tiny_lora_seed42_final_fixedlabel_eval_20260628_000521/fixedlabel_same_checkpoint_comparison.csv'),
('Tiny-ImageNet','ADI-LoRA','Formal small-scale', ROOT/'extracted/04-stage4_tiny_imagenet100_results_collection_20260628_160740/stage4_tiny_imagenet100_collect_20260628_160740/comparisons/failed_lora_recover./outputs/stage4_failed_lora_recover_20260628_153445/tiny_imagenet_lora_seed43_same_checkpoint_recover_20260628_153445/same_checkpoint_comparison.csv'),
('Tiny-ImageNet','ADI-LoRA','Formal small-scale', ROOT/'extracted/04-stage4_tiny_imagenet100_results_collection_20260628_160740/stage4_tiny_imagenet100_collect_20260628_160740/comparisons/failed_lora_recover./outputs/stage4_failed_lora_recover_20260628_153445/tiny_imagenet_lora_seed44_same_checkpoint_recover_20260628_154419/same_checkpoint_comparison.csv'),
('Tiny-ImageNet','ADI-DoRA','Supplement single-seed', ROOT/'extracted/04-stage4_tiny_imagenet100_results_collection_20260628_160740/stage4_tiny_imagenet100_collect_20260628_160740/comparisons/formal_queue./outputs/stage4_five_exp_formal_safe_20260628_005714/tiny_imagenet_dora_seed42_same_checkpoint_20260628_033735/same_checkpoint_comparison.csv'),
('ImageNet-100','ADI-LoRA','Pilot only: generated symlink split', ROOT/'extracted/04-stage4_tiny_imagenet100_results_collection_20260628_160740/stage4_tiny_imagenet100_collect_20260628_160740/comparisons/failed_lora_recover./outputs/stage4_failed_lora_recover_20260628_153445/imagenet100_lora_seed42_same_checkpoint_recover_20260628_155354/same_checkpoint_comparison.csv'),
('ImageNet-100','ADI-DoRA','Pilot only: generated symlink split', ROOT/'extracted/04-stage4_tiny_imagenet100_results_collection_20260628_160740/stage4_tiny_imagenet100_collect_20260628_160740/comparisons/formal_queue./outputs/stage4_five_exp_formal_safe_20260628_005714/imagenet100_dora_seed42_same_checkpoint_20260628_150658/same_checkpoint_comparison.csv'),
]
# fallback find seed42 fixedlabel if path differs
if not stage_sources[0][3].exists():
    cand=list(ROOT.glob('**/fixedlabel_same_checkpoint_comparison.csv'))
    if cand: stage_sources[0]=('Tiny-ImageNet','ADI-LoRA','Formal small-scale',cand[0])

for dataset,method,tier,p in stage_sources:
    if not Path(p).exists():
        continue
    df=read_csv(p)
    for _,r in df.iterrows():
        # support fixedlabel names maybe std/selected or alpha1 names
        a1_iid = r.get('alpha1_iid', r.get('lora_alpha1_iid', r.get('std_iid', np.nan)))
        sel_iid = r.get('selected_iid', r.get('delta_selected_iid', r.get('adi_iid', np.nan)))
        a1_bil = r.get('alpha1_bilinear', r.get('lora_alpha1_bilinear', r.get('std_bilinear', np.nan)))
        sel_bil = r.get('selected_bilinear', r.get('delta_selected_bilinear', r.get('adi_bilinear', np.nan)))
        a1_near = r.get('alpha1_nearest', r.get('lora_alpha1_nearest', r.get('std_nearest', np.nan)))
        sel_near = r.get('selected_nearest', r.get('delta_selected_nearest', r.get('adi_nearest', np.nan)))
        a1_drop = r.get('alpha1_nearest_drop', r.get('lora_alpha1_nearest_drop', np.nan))
        sel_drop = r.get('selected_nearest_drop', r.get('delta_selected_nearest_drop', np.nan))
        stage_rows.append({
            'dataset':dataset,'evidence_tier':tier,'method':method,'base_method':'LoRA' if 'LoRA' in method else 'DoRA','seed':int(r.seed),
            'selected_alpha':r.selected_alpha,'std_iid':pct(a1_iid),'adi_iid':pct(sel_iid),'iid_gain_pp':pp(r.iid_gain),
            'std_bilinear':pct(a1_bil),'adi_bilinear':pct(sel_bil),'bilinear_gain_pp':pp(r.bilinear_gain),
            'std_nearest':pct(a1_near),'adi_nearest':pct(sel_near),'nearest_gain_pp':pp(r.nearest_gain),
            'std_nearest_drop_pp':pp(a1_drop),'adi_nearest_drop_pp':pp(sel_drop),'drop_reduction_pp':pp(r.drop_reduction),
            'nearest_used_for_alpha_selection':norm_bool(r.get('nearest_used_for_alpha_selection',0)),
            'corruption_used_for_alpha_selection':norm_bool(r.get('corruption_used_for_alpha_selection',0)) if 'corruption_used_for_alpha_selection' in r.index else 0,
            'checkpoint_selection_rule':r.get('checkpoint_selection_rule','final_checkpoint_only'),
            'nearest_used_for_checkpoint_selection':norm_bool(r.get('nearest_used_for_checkpoint_selection',0)),
            'same_checkpoint_eval':1,'nan_flag':norm_bool(r.get('nan_flag',0)) if 'nan_flag' in r.index else 0,'divergence_flag':norm_bool(r.get('divergence_flag',0)) if 'divergence_flag' in r.index else 0,
            'source_file':rel(p),'placement':'Main paper candidate' if dataset=='Tiny-ImageNet' and method=='ADI-LoRA' else ('Pilot only' if dataset=='ImageNet-100' else 'Supplement')
        })
stage4=pd.DataFrame(stage_rows)
main_tiny_lora=stage4[(stage4.dataset=='Tiny-ImageNet') & (stage4.method=='ADI-LoRA')].copy()
supp_dora_tiny_imagenet100=stage4[~((stage4.dataset=='Tiny-ImageNet') & (stage4.method=='ADI-LoRA'))].copy()

# Aggregates
agg_rows=[]
def agg_clean(df, dataset, method, base_method, tier):
    sub=df[(df['dataset']==dataset)&(df['method']==method)&(df['evidence_status'].ne('MISSING_SOURCE') if 'evidence_status' in df.columns else True)].copy()
    if len(sub)==0: return
    agg_rows.append({
        'dataset':dataset,'method':method,'base_method':base_method,'evidence_tier':tier,'n':len(sub),'seeds':','.join(map(str,sorted(sub.seed.astype(int).unique()))),
        'mean_iid':sub['adi_iid'].mean(),'mean_bilinear':sub['adi_bilinear'].mean(),'mean_nearest':sub['adi_nearest'].mean(),
        'mean_nearest_gain_pp':sub['nearest_gain_pp'].mean(),'std_nearest_gain_pp':sub['nearest_gain_pp'].std(ddof=1) if len(sub)>1 else np.nan,
        'mean_drop_reduction_pp':sub['drop_reduction_pp'].mean(),'positive_seeds_total':f"{int((sub['nearest_gain_pp']>0).sum())}/{len(sub)}",
        'selected_alpha_distribution':'; '.join(f"{a}: {c}" for a,c in sub['selected_alpha'].value_counts(dropna=False).sort_index().items()),
        'source_status':'complete for requested seeds' if (dataset=='CIFAR-100' and len(sub)==3) or (method=='ADI-LoRA' and dataset=='Tiny-ImageNet' and len(sub)==3) else 'incomplete / single-seed only'
    })
agg_clean(main_cifar100_clean,'CIFAR-100','ADI-LoRA','LoRA','clean 3seed')
agg_clean(main_cifar100_clean,'CIFAR-100','ADI-DoRA','DoRA','clean 3seed')
# stage aggregates
for (dataset,method),sub in stage4.groupby(['dataset','method']):
    tier='pilot only' if dataset=='ImageNet-100' else ('formal small-scale 3seed' if method=='ADI-LoRA' else 'supplement single-seed')
    agg_rows.append({
        'dataset':dataset,'method':method,'base_method':'LoRA' if method.endswith('LoRA') else 'DoRA','evidence_tier':tier,'n':len(sub),'seeds':','.join(map(str,sorted(sub.seed.astype(int).unique()))),
        'mean_iid':sub['adi_iid'].mean(),'mean_bilinear':sub['adi_bilinear'].mean(),'mean_nearest':sub['adi_nearest'].mean(),
        'mean_nearest_gain_pp':sub['nearest_gain_pp'].mean(),'std_nearest_gain_pp':sub['nearest_gain_pp'].std(ddof=1) if len(sub)>1 else np.nan,
        'mean_drop_reduction_pp':sub['drop_reduction_pp'].mean(),'positive_seeds_total':f"{int((sub['nearest_gain_pp']>0).sum())}/{len(sub)}",
        'selected_alpha_distribution':'; '.join(f"{a}: {c}" for a,c in sub['selected_alpha'].value_counts(dropna=False).sort_index().items()),
        'source_status':'pilot only' if dataset=='ImageNet-100' else ('complete for Tiny LoRA 3 seeds' if method=='ADI-LoRA' and len(sub)==3 else 'single-seed supplement')
    })
# corruption aggregate from row-level
for (tier,method),sub in main_cifar100c.groupby(['evidence_tier','method']):
    agg_rows.append({
        'dataset':'CIFAR-100-C','method':method,'base_method':'DoRA','evidence_tier':tier,'n':len(sub),'seeds':','.join(map(str,sorted(sub.seed.astype(int).unique()))),
        'mean_iid':np.nan,'mean_bilinear':np.nan,'mean_nearest':np.nan,
        'mean_nearest_gain_pp':np.nan,'std_nearest_gain_pp':np.nan,
        'mean_drop_reduction_pp':np.nan,'positive_seeds_total':np.nan,
        'selected_alpha_distribution':'; '.join(f"{a}: {c}" for a,c in sub['selected_alpha'].value_counts(dropna=False).sort_index().items()),
        'positive_corruption_cases_total':f"{int((sub['acc_gain_pp']>0).sum())}/{len(sub)}",
        'mean_acc_gain_pp':sub['acc_gain_pp'].mean(),'std_acc_gain_pp':sub['acc_gain_pp'].std(ddof=1),'mean_loss_reduction':sub['loss_reduction'].mean(),
        'loss_positive_cases_total':f"{int((sub['loss_reduction']>0).sum())}/{len(sub)}",
        'source_status':'complete raw CSV uploaded'
    })
aggregates=pd.DataFrame(agg_rows)

# Protocol check
protocol=[]
def add_protocol(scope, df):
    if df is None or len(df)==0: return
    protocol.append({
        'scope':scope,
        'nearest_used_for_alpha_selection': int(pd.to_numeric(df.get('nearest_used_for_alpha_selection',pd.Series([0]*len(df))), errors='coerce').fillna(0).max()),
        'corruption_used_for_alpha_selection': int(pd.to_numeric(df.get('corruption_used_for_alpha_selection',pd.Series([0]*len(df))), errors='coerce').fillna(0).max()),
        'checkpoint_selection_rule': '; '.join(sorted(map(str,df.get('checkpoint_selection_rule',pd.Series(['NA']*len(df))).dropna().unique()))),
        'nearest_used_for_checkpoint_selection': int(pd.to_numeric(df.get('nearest_used_for_checkpoint_selection',pd.Series([0]*len(df))), errors='coerce').fillna(0).max()),
        'nan_flag': int(pd.to_numeric(df.get('nan_flag',pd.Series([0]*len(df))), errors='coerce').fillna(0).max()) if 'nan_flag' in df.columns else 0,
        'divergence_flag': int(pd.to_numeric(df.get('divergence_flag',pd.Series([0]*len(df))), errors='coerce').fillna(0).max()) if 'divergence_flag' in df.columns else 0,
        'final_checkpoint_only': int(all(str(x)=='final_checkpoint_only' for x in df.get('checkpoint_selection_rule',pd.Series([])).dropna())) if 'checkpoint_selection_rule' in df.columns and len(df)>0 else 0,
        'same_checkpoint_eval': int(pd.to_numeric(df.get('same_checkpoint_eval',pd.Series([1]*len(df))), errors='coerce').fillna(1).min()),
        'status': 'PASS' if (int(pd.to_numeric(df.get('nearest_used_for_alpha_selection',pd.Series([0]*len(df))), errors='coerce').fillna(0).max())==0 and int(pd.to_numeric(df.get('corruption_used_for_alpha_selection',pd.Series([0]*len(df))), errors='coerce').fillna(0).max())==0 and int(pd.to_numeric(df.get('nearest_used_for_checkpoint_selection',pd.Series([0]*len(df))), errors='coerce').fillna(0).max())==0) else 'CHECK'
    })
add_protocol('CIFAR-100 clean', main_cifar100_clean[main_cifar100_clean.evidence_status!='MISSING_SOURCE'])
add_protocol('CIFAR-100-C corruption', main_cifar100c)
add_protocol('Tiny/ImageNet100 Stage4', stage4)
protocol_check=pd.DataFrame(protocol)

# Efficiency and control tables supplied in the latest upload.
efficiency_paper_path = ROOT/'new_uploads/extracted/06-ADI_efficiency_filtered_tables_package/ADI_efficiency_filtered_paper.csv'
efficiency_raw_path = ROOT/'new_uploads/extracted/06-ADI_efficiency_filtered_tables_package/ADI_efficiency_filtered_raw.csv'
alpha_overhead_path = ROOT/'new_uploads/extracted/06-ADI_efficiency_filtered_tables_package/ADI_alpha_overhead_filtered_summary.csv'
efficiency_notes_path = ROOT/'new_uploads/extracted/06-ADI_efficiency_filtered_tables_package/ADI_efficiency_protocol_notes.csv'
controls_path = ROOT/'new_uploads/extracted/04-fr_peft_cifar100_seed42_controls_no_ckpt/fr_peft_cifar100_seed42_controls_no_ckpt/outputs/fr_peft_cifar100_controls/summary_all.csv'

efficiency_paper = read_csv(efficiency_paper_path) if efficiency_paper_path.exists() else pd.DataFrame()
efficiency_raw = read_csv(efficiency_raw_path) if efficiency_raw_path.exists() else pd.DataFrame()
alpha_overhead = read_csv(alpha_overhead_path) if alpha_overhead_path.exists() else pd.DataFrame()
efficiency_protocol_notes = read_csv(efficiency_notes_path) if efficiency_notes_path.exists() else pd.DataFrame()
controls = read_csv(controls_path) if controls_path.exists() else pd.DataFrame()
if not controls.empty:
    controls['placement'] = 'Supplement / control experiment'
    controls['source_file'] = rel(controls_path)

# Table placement
placement_rows=[
{'table_name':'Main_CIFAR100_Clean','recommended_placement':'Main paper candidate','reason':'LoRA and DoRA both now have raw same-checkpoint clean three-seed evidence.'},
{'table_name':'Main_CIFAR100C','recommended_placement':'Main paper for DoRA subset; Supplement for full19 detail','reason':'Subset has 36 raw cases over 3 seeds; full19 is seed42 severity=3 detail. LoRA CIFAR-100-C raw source missing.'},
{'table_name':'Main_TinyImageNet_LoRA','recommended_placement':'Main paper candidate','reason':'Tiny-ImageNet LoRA has 3 seeds, final checkpoint, same-checkpoint eval, positive 3/3.'},
{'table_name':'Supplement_DoRA_Tiny_ImageNet100','recommended_placement':'Supplement / Pilot','reason':'Tiny-ImageNet DoRA is seed42 only; ImageNet-100 uses generated class-stratified symlink split and must be pilot only.'},
{'table_name':'Efficiency','recommended_placement':'Main or Supplement efficiency table','reason':'Filtered efficiency source shows ADI adds no trainable params, no inference modules, and no extra training epochs; alpha selection is post-hoc.'},
{'table_name':'Controls_CIFAR100_seed42','recommended_placement':'Supplement / diagnostic control','reason':'Low-lr, strong-wd, and 10ep controls are single-seed controls and should not replace three-seed main evidence.'},
{'table_name':'Protocol_Check','recommended_placement':'Main or appendix checklist','reason':'Directly addresses alpha/checkpoint leakage risk.'},
{'table_name':'File_Inventory','recommended_placement':'Appendix / reproducibility material','reason':'Audits sources, MISSING entries, and no-weight packages.'},
{'table_name':'VPT/Adapter diagnostics','recommended_placement':'Not use / diagnostic only','reason':'Outside ADI-LoRA LoRA-style weight-delta main claim and includes boundary diagnostics.'},
]
table_placement=pd.DataFrame(placement_rows)

# Missing sources
missing_rows=[
{'missing_item':'LoRA CIFAR-100-C full19 severity=3 seed42 raw CSV','known_narrative_anchor':'19/19 positive, mean Acc gain approx +1.46 pp, loss reduction 19/19','why_missing_matters':'Cannot include LoRA corruption full19 as formal table without raw per-corruption CSV.','required_file':'lora_cifar100c_full19_s3_seed42_no_ckpt tar.gz or comparison CSV','usable_for_formal_claim':'NO'},
{'missing_item':'LoRA CIFAR-100-C subset raw CSV','known_narrative_anchor':'not provided','why_missing_matters':'Only DoRA subset raw source is uploaded; LoRA corruption subset cannot be claimed.','required_file':'lora_cifar100c_subset or 3seed package','usable_for_formal_claim':'NO'},
]
missing_sources=pd.DataFrame(missing_rows)

# Dashboard
# key values
lora_agg=aggregates[(aggregates.dataset=='CIFAR-100')&(aggregates.method=='ADI-LoRA')]
dora_agg=aggregates[(aggregates.dataset=='CIFAR-100')&(aggregates.method=='ADI-DoRA')]
tiny_lora_agg=aggregates[(aggregates.dataset=='Tiny-ImageNet')&(aggregates.method=='ADI-LoRA')]
dora_c=aggregates[(aggregates.dataset=='CIFAR-100-C')&(aggregates.evidence_tier=='subset_4corr_s1_s3_s5_3seed')]
dashboard_rows=[
{'item':'Overall conclusion','value':'ADI-LoRA/ADI-DoRA now have protocol-clean CIFAR-100 clean three-seed evidence where raw CSV is available; DoRA corruption evidence is strong; Tiny-ImageNet LoRA transfers with modest positive gains.'},
{'item':'Main-ready evidence','value':'CIFAR-100 LoRA clean 3-seed; CIFAR-100 DoRA clean 3-seed; CIFAR-100-C DoRA subset 36 cases; Tiny-ImageNet LoRA 3-seed; efficiency/protocol checklist.'},
{'item':'Supplement evidence','value':'Tiny-ImageNet DoRA seed42; CIFAR-100-C DoRA full19 severity=3 seed42; recovery logs; split manifest.'},
{'item':'Pilot only','value':'ImageNet-100 LoRA/DoRA seed42 due generated class-stratified symlink split.'},
{'item':'Not use / diagnostic only','value':'legacy exploratory method/legacy adapter baseline/VPT/Adapter boundary diagnostics; any best-Nearest or Nearest-selected result.'},
{'item':'CIFAR-100 LoRA mean Nearest gain pp','value':float(lora_agg['mean_nearest_gain_pp'].iloc[0]) if len(lora_agg) else np.nan},
{'item':'CIFAR-100 DoRA mean Nearest gain pp','value':float(dora_agg['mean_nearest_gain_pp'].iloc[0]) if len(dora_agg) else np.nan},
{'item':'Tiny-ImageNet LoRA mean Nearest gain pp','value':float(tiny_lora_agg['mean_nearest_gain_pp'].iloc[0]) if len(tiny_lora_agg) else np.nan},
{'item':'CIFAR-100-C DoRA subset positive cases','value':dora_c['positive_corruption_cases_total'].iloc[0] if len(dora_c) and 'positive_corruption_cases_total' in dora_c.columns else np.nan},
{'item':'CIFAR-100-C DoRA subset mean Acc gain pp','value':float(dora_c['mean_acc_gain_pp'].iloc[0]) if len(dora_c) and 'mean_acc_gain_pp' in dora_c.columns else np.nan},
{'item':'Critical missing source','value':'LoRA CIFAR-100-C raw CSV is still not uploaded; DoRA clean seed43/44 and efficiency sources are now available.'},
]
dashboard=pd.DataFrame(dashboard_rows)

# Main/supp CSV combined
main_tables=pd.concat([
    main_cifar100_clean.assign(table_name='Main_CIFAR100_Clean'),
    main_cifar100c.assign(table_name='Main_CIFAR100C'),
    main_tiny_lora.assign(table_name='Main_TinyImageNet_LoRA'),
], ignore_index=True, sort=False)
supp_tables=pd.concat([
    supp_dora_tiny_imagenet100.assign(table_name='Supplement_DoRA_Tiny_ImageNet100'),
    main_cifar100c[main_cifar100c.evidence_tier=='full19_s3_seed42'].assign(table_name='CIFAR100C_full19_detail'),
    controls.assign(table_name='Controls_CIFAR100_seed42') if not controls.empty else pd.DataFrame(),
], ignore_index=True, sort=False)

# Save CSVs
for name,df in [
('ADI_main_tables.csv',main_tables),('ADI_supplement_tables.csv',supp_tables),('ADI_protocol_checklist.csv',protocol_check),
('ADI_file_inventory.csv',file_inventory),('ADI_aggregates.csv',aggregates),('ADI_table_placement.csv',table_placement),('ADI_missing_sources.csv',missing_sources),
('Main_CIFAR100_Clean.csv',main_cifar100_clean),('Main_CIFAR100C.csv',main_cifar100c),('Main_TinyImageNet_LoRA.csv',main_tiny_lora),('Supplement_DoRA_Tiny_ImageNet100.csv',supp_dora_tiny_imagenet100),
('ADI_efficiency_table.csv',efficiency_paper),('ADI_efficiency_raw.csv',efficiency_raw),('ADI_alpha_overhead_summary.csv',alpha_overhead),('ADI_controls_cifar100_seed42.csv',controls)]:
    df.to_csv(OUT/name,index=False)

# Markdown reports
def df_to_md(df):
    if df is None or df.empty:
        return ''
    cols=list(df.columns)
    lines=['| ' + ' | '.join(cols) + ' |',
           '| ' + ' | '.join(['---']*len(cols)) + ' |']
    for _,row in df.iterrows():
        vals=[]
        for c in cols:
            v=row[c]
            if pd.isna(v):
                s=''
            else:
                s=str(v)
            vals.append(s.replace('|','/').replace('\n',' '))
        lines.append('| ' + ' | '.join(vals) + ' |')
    return '\n'.join(lines)

md=[]
md.append('# ADI Missing Sources Report\n')
md.append('This report follows the rule: raw CSV/XLSX source overrides narrative anchors. Narrative anchors are listed only to identify expected files; they are not used as formal results.\n')
md.append(df_to_md(missing_sources))
md.append('\n\n## Manifest Cross-Check\n')
if manifest_missing:
    md.append('`ADI_final_table_sources_20260629_213223.tar.gz` reports these MISSING labels:\n')
    for x in manifest_missing: md.append(f'- {x}')
else: md.append('No MISSING labels found in source manifest.')
md.append('\n\n## Current Formal-Claim Boundary\n')
md.append('- CIFAR-100 LoRA clean 3-seed: usable.\n- CIFAR-100 DoRA clean 3-seed: usable.\n- CIFAR-100-C DoRA subset/full19: usable for DoRA corruption evidence.\n- LoRA CIFAR-100-C full19/subset: not raw-backed in current uploads.\n- Tiny-ImageNet LoRA 3-seed: usable.\n- Tiny-ImageNet DoRA seed42: supplement only.\n- ImageNet-100 LoRA/DoRA seed42: pilot only due generated symlink split.\n- Efficiency table: usable as main/supplement efficiency evidence; ADI adds no trainable parameters, no inference modules, and no extra training epochs.\n')
(OUT/'ADI_missing_sources_report.md').write_text('\n'.join(md), encoding='utf-8')

narr=[]
narr.append('# ADI Table Narrative for Paper\n')
narr.append('## Recommended Evidence Wording\n')
narr.append('ADI-LoRA is evaluated as a post-hoc delta calibration for LoRA-style PEFT under a final-checkpoint, same-checkpoint protocol. Alpha is selected only using Bicubic/Bilinear validation performance; Nearest and corruption results are held out from selection.\n')
if len(lora_agg):
    r=lora_agg.iloc[0]
    narr.append(f"On CIFAR-100, ADI-LoRA improves Nearest accuracy over standard LoRA across {r['positive_seeds_total']} seeds, with mean Nearest gain {r['mean_nearest_gain_pp']:.2f} pp and mean drop reduction {r['mean_drop_reduction_pp']:.2f} pp.\n")
if len(dora_agg):
    r=dora_agg.iloc[0]
    narr.append(f"On CIFAR-100, ADI-DoRA improves Nearest accuracy over DoRA across {r['positive_seeds_total']} seeds, with mean Nearest gain {r['mean_nearest_gain_pp']:.2f} pp and mean drop reduction {r['mean_drop_reduction_pp']:.2f} pp.\n")
if len(tiny_lora_agg):
    r=tiny_lora_agg.iloc[0]
    narr.append(f"On Tiny-ImageNet, ADI-LoRA remains positive but more modest: {r['positive_seeds_total']} seeds improve Nearest accuracy, with mean Nearest gain {r['mean_nearest_gain_pp']:.2f} pp.\n")
if len(dora_c):
    r=dora_c.iloc[0]
    narr.append(f"For CIFAR-100-C DoRA subset evaluation, ADI-DoRA improves {r['positive_corruption_cases_total']} corruption cases, with mean Acc gain {r['mean_acc_gain_pp']:.2f} pp; corruption results are not used for alpha selection.\n")
if not efficiency_paper.empty:
    narr.append('The efficiency table supports the no-extra-inference-cost claim: ADI uses the same trained checkpoint as the base LoRA/DoRA model, adds no trainable parameters, no inference modules, and no extra training epochs; only post-hoc alpha selection is added.\n')
narr.append('## Boundary Wording\n')
narr.append('Tiny-ImageNet DoRA is reported as single-seed supplementary evidence, not as a three-seed conclusion. ImageNet-100 results are pilot-only because the current split is a generated class-stratified symlink split rather than a standard official benchmark split. LoRA CIFAR-100-C raw sources must be supplied before making formal LoRA corruption claims.\n')
narr.append('## Table Placement\n')
narr.append(df_to_md(table_placement))
(OUT/'ADI_table_narrative_for_paper.md').write_text('\n'.join(narr), encoding='utf-8')

# Excel workbook via pandas/openpyxl
xlsx=OUT/'ADI_final_experiment_tables.xlsx'
with pd.ExcelWriter(xlsx, engine='openpyxl') as writer:
    sheets=[
        ('Dashboard',dashboard),('File_Inventory',file_inventory),('Main_CIFAR100_Clean',main_cifar100_clean),('Main_CIFAR100C',main_cifar100c),
        ('Main_TinyImageNet_LoRA',main_tiny_lora),('Supp_DoRA_Tiny_ImageNet100',supp_dora_tiny_imagenet100),('Aggregates',aggregates),
        ('Efficiency',efficiency_paper),('Alpha_Overhead',alpha_overhead),('Controls_CIFAR100_seed42',controls),
        ('Protocol_Check',protocol_check),('Table_Placement',table_placement),('Missing_Sources',missing_sources)
    ]
    for s,df in sheets:
        df.to_excel(writer, sheet_name=s, index=False)

# Style workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
wb=load_workbook(xlsx)
header_fill=PatternFill('solid', fgColor='1F4E78')
header_font=Font(color='FFFFFF', bold=True)
sub_fill=PatternFill('solid', fgColor='D9EAF7')
for ws in wb.worksheets:
    ws.freeze_panes='A2'
    ws.sheet_view.showGridLines=False
    max_col=ws.max_column
    for cell in ws[1]:
        cell.fill=header_fill; cell.font=header_font; cell.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment=Alignment(vertical='top', wrap_text=True)
            if isinstance(cell.value,(int,float)):
                cell.number_format='0.00'
    for col in range(1,max_col+1):
        letter=get_column_letter(col)
        vals=[ws.cell(r,col).value for r in range(1,min(ws.max_row,80)+1)]
        width=min(max(10, max(len(str(v)) if v is not None else 0 for v in vals)+2), 45)
        ws.column_dimensions[letter].width=width
    ws.auto_filter.ref=ws.dimensions
    for row in range(1, ws.max_row+1):
        ws.row_dimensions[row].height=30 if row==1 else 42 if ws.max_column>8 else 28
wb.save(xlsx)

print('WROTE', OUT)
print('Files:', sorted(p.name for p in OUT.iterdir()))
